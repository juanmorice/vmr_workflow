from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = 'Requests'

# Headers
headers = [
    'ID', 'Client Name', 'Program Name', 'Reward BL', 'Announcement BL',
    'Promoted Groups', 'Category Groups', 'Threshold Metric', 'Segment Definitions',
    'Requested by', 'Email', 'LMC List ID', 'Retailer Category Level', 'Mapping File name',
    'Status', 'Created Date'
]

# Write headers with styling
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Status column styling (read-only appearance - gray background)
status_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
id_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

# Add auto-incrementing ID and UNDONE status to all rows (500 rows for data entry)
for row in range(2, 502):
    ws.cell(row=row, column=1, value=row - 1)  # ID column (A) - auto-assigned
    ws.cell(row=row, column=1).fill = id_fill
    ws.cell(row=row, column=15, value='UNDONE')  # Status column (O) - before Created Date
    ws.cell(row=row, column=15).fill = status_fill
    # Created Date formula - auto-populates when Client Name is entered
    ws.cell(row=row, column=16, value=f'=IF(B{row}<>"",NOW(),"")') 
    ws.cell(row=row, column=16).fill = id_fill
    ws.cell(row=row, column=16).number_format = 'YYYY-MM-DD HH:MM:SS'

# Data validation for Threshold Metric (column H)
threshold_dv = DataValidation(
    type='list',
    formula1='"units,dollars"',
    allow_blank=True
)
threshold_dv.error = 'Please select units or dollars'
threshold_dv.errorTitle = 'Invalid Threshold'
ws.add_data_validation(threshold_dv)
threshold_dv.add('H2:H501')

# Data validation for Segment Definitions (column I)
segment_dv = DataValidation(
    type='list',
    formula1='"1 - UPC list groups,2 - Catalina\'s brand description,3 - Catalina\'s category description,4 - Retailer\'s own descriptions (except Kroger),5 - Kroger\'s own descriptions,6 - Custom Brand\'s descriptions (from provided file)"',
    allow_blank=True
)
segment_dv.error = 'Please select a valid segment definition'
segment_dv.errorTitle = 'Invalid Segment'
ws.add_data_validation(segment_dv)
segment_dv.add('I2:I501')

# Data validation for Retailer Category Level (column M)
category_dv = DataValidation(
    type='list',
    formula1='"Level 1,Level 2,Level 3,Level 4,Level 5"',
    allow_blank=True
)
category_dv.error = 'Please select a valid category level'
category_dv.errorTitle = 'Invalid Category Level'
ws.add_data_validation(category_dv)
category_dv.add('M2:M501')

# Data validation for Status (column O) - locked to UNDONE for new entries
status_dv = DataValidation(
    type='list',
    formula1='"UNDONE"',
    allow_blank=False
)
status_dv.error = 'Status cannot be modified manually'
status_dv.errorTitle = 'Read-Only Field'
ws.add_data_validation(status_dv)
status_dv.add('O2:O501')

# Adjust column widths
col_widths = {
    'A': 8, 'B': 18, 'C': 22, 'D': 20, 'E': 18, 'F': 18, 'G': 18,
    'H': 18, 'I': 35, 'J': 20, 'K': 30, 'L': 38, 'M': 22, 'N': 20, 'O': 12, 'P': 20
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

wb.save('src/data_template.xlsx')
print('Excel template with dropdowns created: src/data_template.xlsx')
print('Extended to 500 rows with pre-populated IDs and Status')
print('Column order: ... Email | LMC List ID | Retailer Category Level | Mapping File name | Status | Created Date')
